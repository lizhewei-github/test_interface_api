# encoding:utf-8
import yaml
from openpyxl import load_workbook


class Read_Data:
    def get_yaml_data(self, yaml_path=None):
        '''读取yaml文件'''
        try:
            f = open(yaml_path, "r", encoding="utf-8")
            yaml_data = yaml.load_all(f, Loader=yaml.FullLoader)
            return {
                "msg": "success",
                "result": list(yaml_data)
            }
        except Exception as e:
            print(e)

    def get_txt_data(self, txt_path=None):
        '''读取txt文件'''
        try:
            with open(txt_path, "r") as f:
                txt_data = f.readlines()
            print(txt_data)
        except Exception as e:
            print(e)

    def get_excel_data(self, excel_path, sheet_name):
        '''获取excel表格'''
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]
        test_data = []
        for i in range(2, sheet.max_row + 1):  # 最大列
            row_data = {}
            # 开始循环第二行.第三行，第四行，第五行。。。。。。1列
            row_data["case_id"] = sheet.cell(i, 1).value
            row_data["case_name"] = sheet.cell(i, 2).value
            row_data["url"] = sheet.cell(i, 3).value
            row_data["method"] = sheet.cell(i, 4).value
            row_data["headers"] = sheet.cell(i, 5).value
            row_data["params"] = sheet.cell(i, 6).value
            row_data["checkpoint"] = sheet.cell(i, 7).value
            test_data.append(row_data)
        return test_data
        # print(test_data)


'''
        frinal_data=[]
        if button=='off':
            frinal_data=test_data
        else:
            for item in test_data:
                #进行循环，iten列表循环id，列表为配置文件，eval格式化，去皮
                if item["case_id"] in eval(button):
                    frinal_data.append(item)
        return frinal_data
'''

if __name__ == '__main__':
    ss = Read_Data()
    ss.get_yaml_data(yaml_path='../config/config.yaml')
